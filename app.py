import ast
import io
import json
import os
import sys
import smtplib
import subprocess
import tempfile
import threading
from concurrent.futures import ThreadPoolExecutor
from functools import wraps
from collections import OrderedDict
from datetime import date

import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from flask import Flask, flash, redirect, render_template, request, session, url_for
import sqlite3

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'change-me-in-production')

DB_PATH           = os.environ.get('DB_PATH', 'database.db')
ADMIN_PASSWORD    = os.environ.get('ADMIN_PASSWORD', 'admin123')
PROFESSOR_EMAIL   = os.environ.get('PROFESSOR_EMAIL', '')
SMTP_HOST         = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT         = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER         = os.environ.get('SMTP_USER', '')
SMTP_PASS         = os.environ.get('SMTP_PASS', '')
SENDER_NAME       = os.environ.get('SENDER_NAME', 'Student Platform')
ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')


# ─── Database ──────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS questions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            description TEXT NOT NULL,
            template TEXT DEFAULT '',
            order_num INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS submissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_email TEXT NOT NULL,
            submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            sent INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS answers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            submission_id INTEGER NOT NULL,
            question_id INTEGER NOT NULL,
            answer_text TEXT DEFAULT '',
            claude_feedback TEXT DEFAULT '',
            FOREIGN KEY (submission_id) REFERENCES submissions(id),
            FOREIGN KEY (question_id) REFERENCES questions(id)
        );
    ''')
    for sql in [
        'ALTER TABLE questions ADD COLUMN template TEXT DEFAULT ""',
        'ALTER TABLE answers ADD COLUMN claude_feedback TEXT DEFAULT ""',
    ]:
        try:
            conn.execute(sql)
        except Exception:
            pass
    conn.commit()
    conn.close()


# ─── Auth ──────────────────────────────────────────────────────────────────────

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated


# ─── Code utilities ────────────────────────────────────────────────────────────

def _claude():
    return anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)


def _find_fn_name(student_code, template=''):
    def first_fn(code):
        try:
            for node in ast.walk(ast.parse(code)):
                if isinstance(node, ast.FunctionDef):
                    return node.name
        except Exception:
            pass
        return None

    if template:
        tpl_fn = first_fn(template)
        if tpl_fn:
            try:
                for node in ast.walk(ast.parse(student_code)):
                    if isinstance(node, ast.FunctionDef) and node.name == tpl_fn:
                        return tpl_fn
            except Exception:
                pass
    return first_fn(student_code)


def run_student_code(code, timeout=5):
    try:
        r = subprocess.run(
            [sys.executable, '-c', code],
            capture_output=True, text=True, timeout=timeout
        )
        return r.stdout, r.stderr, r.returncode
    except subprocess.TimeoutExpired:
        return '', f'Execution timed out after {timeout}s.', -1
    except Exception as exc:
        return '', str(exc), -1


# ─── Claude: safety check ──────────────────────────────────────────────────────

def check_code_safety(code):
    if not ANTHROPIC_API_KEY:
        return True, 'No API key — safety check skipped'
    prompt = (
        'You are a code security checker. Analyze this Python code ONLY for safety risks.\n\n'
        'Flag if you see:\n'
        '- Dangerous imports: os.system, subprocess, shutil.rmtree, open() for writing, '
        'socket, requests, urllib, eval, exec, __import__\n'
        '- Infinite loops (while True with no break, unbounded recursion)\n'
        '- File system writes, network calls, or shell commands\n\n'
        'Reply with ONLY one line:\n'
        'SAFE\n'
        'or\n'
        'UNSAFE: <brief reason, max 20 words>\n\n'
        f'Code:\n```python\n{code}\n```'
    )
    try:
        resp = _claude().messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=60,
            messages=[{'role': 'user', 'content': prompt}]
        )
        line = resp.content[0].text.strip().splitlines()[0]
        if line.upper().startswith('UNSAFE'):
            reason = line[7:].strip() if len(line) > 7 else 'Dangerous code detected'
            return False, reason
        return True, 'OK'
    except Exception as exc:
        return True, f'Safety check error: {exc}'


# ─── Claude: reference solution + test generator ───────────────────────────────

def generate_reference_solution(question, student_code):
    if not ANTHROPIC_API_KEY:
        return None
    prompt = (
        f'Problem: {question["title"]}\n'
        f'{question["description"]}\n\n'
        f'Student code (use only to determine function signature):\n'
        f'```python\n{student_code}\n```\n\n'
        'Write exactly two Python functions — no markdown, no explanation, only valid Python:\n'
        '1. `reference_solution(...)` — correct implementation with same signature as the student.\n'
        '2. `generate_test_inputs(n)` — returns a list of n tuples (args for one call), '
        'covering normal cases, edge cases, and stress inputs.'
    )
    try:
        resp = _claude().messages.create(
            model='claude-sonnet-4-6',
            max_tokens=1024,
            messages=[{'role': 'user', 'content': prompt}]
        )
        code = resp.content[0].text.strip()
        if code.startswith('```'):
            lines = code.splitlines()
            code = '\n'.join(lines[1:-1] if lines[-1].strip() == '```' else lines[1:])
        return code
    except Exception:
        return None


# ─── Test runner ───────────────────────────────────────────────────────────────

def run_correctness_tests(student_code, reference_code, student_fn_name, n=1000, timeout=20):
    parts = [
        'import json, random, math, time, tracemalloc\n\n',
        '# ── Reference solution + test input generator ──\n',
        reference_code, '\n\n',
        '# ── Student solution ──\n',
        student_code, '\n\n',
        'try:\n',
        '    student_fn = ', student_fn_name, '\n',
        '    inputs = generate_test_inputs(', str(n), ')\n',
        '    passed = 0\n',
        '    wrong_answers = 0\n',
        '    exceptions_count = 0\n',
        '    failures = []\n',
        '    times = []\n\n',
        '    tracemalloc.start()\n\n',
        '    for inp in inputs:\n',
        '        if not isinstance(inp, tuple):\n',
        '            inp = (inp,)\n',
        '        try:\n',
        '            ref_out = reference_solution(*inp)\n',
        '        except Exception:\n',
        '            continue\n',
        '        t0 = time.perf_counter()\n',
        '        try:\n',
        '            got = student_fn(*inp)\n',
        '            times.append((time.perf_counter() - t0) * 1000)\n',
        '            if got == ref_out:\n',
        '                passed += 1\n',
        '            else:\n',
        '                wrong_answers += 1\n',
        '                if len(failures) < 3:\n',
        '                    failures.append({"input": repr(inp), "expected": repr(ref_out), "got": repr(got)})\n',
        '        except Exception as e:\n',
        '            times.append((time.perf_counter() - t0) * 1000)\n',
        '            exceptions_count += 1\n',
        '            if len(failures) < 3:\n',
        '                failures.append({"input": repr(inp), "expected": repr(ref_out), "got": "Error: " + str(e)})\n\n',
        '    _, peak = tracemalloc.get_traced_memory()\n',
        '    tracemalloc.stop()\n\n',
        '    total = len(inputs)\n',
        '    avg_t = round(sum(times) / len(times), 3) if times else 0\n',
        '    max_t = round(max(times), 3) if times else 0\n',
        '    min_t = round(min(times), 3) if times else 0\n',
        '    print(json.dumps({\n',
        '        "passed": passed, "total": total,\n',
        '        "wrong_answers": wrong_answers, "exceptions": exceptions_count,\n',
        '        "avg_time_ms": avg_t, "max_time_ms": max_t, "min_time_ms": min_t,\n',
        '        "peak_memory_mb": round(peak / 1024 / 1024, 4),\n',
        '        "failures": failures\n',
        '    }))\n',
        'except Exception as e:\n',
        '    print(json.dumps({"error": str(e), "passed": 0, "total": 0, "failures": []}))\n',
    ]
    script = ''.join(parts)

    with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False, encoding='utf-8') as f:
        f.write(script)
        fname = f.name

    try:
        r = subprocess.run([sys.executable, fname],
                           capture_output=True, text=True, timeout=timeout)
        try:
            os.unlink(fname)
        except Exception:
            pass
        if r.stdout.strip():
            return json.loads(r.stdout.strip())
        return {'error': (r.stderr or 'No output')[:400], 'passed': 0, 'total': n, 'failures': []}
    except subprocess.TimeoutExpired:
        try:
            os.unlink(fname)
        except Exception:
            pass
        return {'error': f'Tests timed out after {timeout}s', 'passed': 0, 'total': n, 'failures': []}
    except Exception as exc:
        try:
            os.unlink(fname)
        except Exception:
            pass
        return {'error': str(exc), 'passed': 0, 'total': n, 'failures': []}


# ─── Analysis (runs on Send Now, per submission) ───────────────────────────────

def _analyze_one_answer(answer_id, answer_text, q):
    """Analyze a single answer. Returns (answer_id, record_dict)."""
    is_safe, safety_reason = check_code_safety(answer_text)
    record = {'safety': 'CLEARED' if is_safe else 'NOT CLEARED', 'safety_reason': safety_reason}

    if not is_safe:
        record['note'] = 'Code not executed due to safety flag'
    else:
        fn_name = _find_fn_name(answer_text, q['template'])
        if not fn_name:
            record['note'] = 'No function definition found'
        else:
            ref_code = generate_reference_solution(q, answer_text)
            if not ref_code:
                record['note'] = 'Could not generate reference solution'
            else:
                tr = run_correctness_tests(answer_text, ref_code, fn_name)
                total  = tr.get('total', 0)
                passed = tr.get('passed', 0)
                record.update({
                    'passed':         passed,
                    'total':          total,
                    'pass_rate':      round(passed / total * 100, 1) if total > 0 else 0,
                    'wrong_answers':  tr.get('wrong_answers', 0),
                    'exceptions':     tr.get('exceptions', 0),
                    'avg_time_ms':    tr.get('avg_time_ms', 0),
                    'max_time_ms':    tr.get('max_time_ms', 0),
                    'min_time_ms':    tr.get('min_time_ms', 0),
                    'peak_memory_mb': tr.get('peak_memory_mb', 0),
                    'failures':       tr.get('failures', []),
                    'error':          tr.get('error', ''),
                })
    return answer_id, record


def _analyze_one_submission(submission_id, questions_list):
    """Analyze all unanalyzed code answers for one submission in parallel."""
    conn = get_db()
    answers = conn.execute(
        'SELECT * FROM answers WHERE submission_id=?', (submission_id,)
    ).fetchall()
    conn.close()

    tasks = []
    for a in answers:
        if a['claude_feedback']:
            continue
        q = next((q for q in questions_list if q['id'] == a['question_id']), None)
        if q and q['template'] and a['answer_text']:
            tasks.append((a['id'], a['answer_text'], q))

    if not tasks:
        return

    # Analyze all questions for this student in parallel
    results = []
    with ThreadPoolExecutor(max_workers=len(tasks)) as ex:
        futures = [ex.submit(_analyze_one_answer, aid, atxt, q) for aid, atxt, q in tasks]
        for f in futures:
            try:
                results.append(f.result())
            except Exception as e:
                print(f'[Analyze answer] {e}')

    # Write all results in one DB transaction
    conn = get_db()
    for answer_id, record in results:
        conn.execute('UPDATE answers SET claude_feedback=? WHERE id=?',
                     (json.dumps(record), answer_id))
    conn.commit()
    conn.close()


# ─── Student routes ────────────────────────────────────────────────────────────

@app.route('/')
def index():
    conn = get_db()
    questions = conn.execute('SELECT * FROM questions ORDER BY order_num, id').fetchall()
    conn.close()
    return render_template('index.html', questions=questions)


@app.route('/submit', methods=['POST'])
def submit():
    student_email = request.form.get('student_email', '').strip()
    if not student_email or '@' not in student_email:
        flash('Please enter a valid email address.', 'error')
        return redirect(url_for('index'))

    send_confirmation = request.form.get('send_confirmation') == 'on'

    conn = get_db()
    questions = conn.execute('SELECT * FROM questions ORDER BY order_num, id').fetchall()
    cur = conn.execute('INSERT INTO submissions (student_email) VALUES (?)', (student_email,))
    submission_id = cur.lastrowid

    questions_list = [dict(q) for q in questions]
    answers_map = {}
    for q in questions_list:
        answer = request.form.get(f'answer_{q["id"]}', '').strip()
        answers_map[q['id']] = answer
        conn.execute(
            'INSERT INTO answers (submission_id, question_id, answer_text) VALUES (?, ?, ?)',
            (submission_id, q['id'], answer)
        )

    conn.commit()
    conn.close()

    if send_confirmation:
        threading.Thread(
            target=_send_student_confirmation,
            args=(student_email, questions_list, answers_map)
        ).start()

    return render_template('success.html', email=student_email)


# ─── Admin routes ──────────────────────────────────────────────────────────────

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        if request.form.get('password') == ADMIN_PASSWORD:
            session['admin'] = True
            return redirect(url_for('admin_dashboard'))
        flash('Incorrect password.', 'error')
    return render_template('admin/login.html')


@app.route('/admin/logout')
def admin_logout():
    session.pop('admin', None)
    return redirect(url_for('admin_login'))


@app.route('/admin')
@admin_required
def admin_dashboard():
    conn = get_db()
    questions = conn.execute('''
        SELECT q.*, COUNT(DISTINCT a.submission_id) AS submission_count
        FROM questions q
        LEFT JOIN answers a ON a.question_id = q.id
        GROUP BY q.id
        ORDER BY q.order_num, q.id
    ''').fetchall()
    total  = conn.execute('SELECT COUNT(*) FROM submissions').fetchone()[0]
    unsent = conn.execute('SELECT COUNT(*) FROM submissions WHERE sent=0').fetchone()[0]

    # Count distinct students with unsent submissions
    unsent_students = conn.execute(
        'SELECT COUNT(DISTINCT student_email) FROM submissions WHERE sent=0'
    ).fetchone()[0]
    conn.close()
    return render_template('admin/dashboard.html', questions=questions,
                           total=total, unsent=unsent, unsent_students=unsent_students)


@app.route('/admin/questions/add', methods=['GET', 'POST'])
@admin_required
def admin_add_question():
    if request.method == 'POST':
        title    = request.form.get('title', '').strip()
        desc     = request.form.get('description', '').strip()
        template = request.form.get('template', '').strip()
        order    = int(request.form.get('order_num', '0') or '0')
        if title and desc:
            conn = get_db()
            conn.execute('INSERT INTO questions (title, description, template, order_num) VALUES (?,?,?,?)',
                         (title, desc, template, order))
            conn.commit()
            conn.close()
            flash('Question added.', 'success')
            return redirect(url_for('admin_dashboard'))
        flash('Title and description are required.', 'error')
    return render_template('admin/question_form.html', question=None, action='Add')


@app.route('/admin/questions/bulk-add', methods=['GET', 'POST'])
@admin_required
def admin_bulk_add():
    if request.method == 'POST':
        indices = request.form.get('indices', '').split(',')
        conn = get_db()
        added = 0
        for i in indices:
            i = i.strip()
            title    = request.form.get(f'title_{i}', '').strip()
            desc     = request.form.get(f'description_{i}', '').strip()
            template = request.form.get(f'template_{i}', '').strip()
            order    = int(request.form.get(f'order_num_{i}', '0') or '0')
            if title and desc:
                conn.execute('INSERT INTO questions (title, description, template, order_num) VALUES (?,?,?,?)',
                             (title, desc, template, order))
                added += 1
        conn.commit()
        conn.close()
        flash(f'{added} question(s) added.', 'success')
        return redirect(url_for('admin_dashboard'))
    return render_template('admin/bulk_add.html')


@app.route('/admin/questions/edit/<int:qid>', methods=['GET', 'POST'])
@admin_required
def admin_edit_question(qid):
    conn = get_db()
    question = conn.execute('SELECT * FROM questions WHERE id=?', (qid,)).fetchone()
    if not question:
        conn.close()
        flash('Question not found.', 'error')
        return redirect(url_for('admin_dashboard'))
    if request.method == 'POST':
        title    = request.form.get('title', '').strip()
        desc     = request.form.get('description', '').strip()
        template = request.form.get('template', '').strip()
        order    = int(request.form.get('order_num', '0') or '0')
        if title and desc:
            conn.execute('UPDATE questions SET title=?, description=?, template=?, order_num=? WHERE id=?',
                         (title, desc, template, order, qid))
            conn.commit()
            conn.close()
            flash('Question updated.', 'success')
            return redirect(url_for('admin_dashboard'))
        flash('Title and description are required.', 'error')
    conn.close()
    return render_template('admin/question_form.html', question=question, action='Edit')


@app.route('/admin/questions/delete/<int:qid>', methods=['POST'])
@admin_required
def admin_delete_question(qid):
    conn = get_db()
    conn.execute('DELETE FROM answers WHERE question_id=?', (qid,))
    conn.execute('DELETE FROM questions WHERE id=?', (qid,))
    conn.commit()
    conn.close()
    flash('Question deleted.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/send-now', methods=['POST'])
@admin_required
def admin_send_now():
    success, message = send_report()
    flash(message, 'success' if success else 'error')
    return redirect(url_for('admin_dashboard'))


# ─── Email helpers ─────────────────────────────────────────────────────────────

def _smtp():
    srv = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
    srv.ehlo(); srv.starttls(); srv.login(SMTP_USER, SMTP_PASS)
    return srv


def _send_student_confirmation(student_email, questions, answers_map):
    if not all([SMTP_USER, SMTP_PASS]):
        return
    try:
        lines = ['Your answers have been received.\n']
        for i, q in enumerate(questions, 1):
            lines.append(f'Q{i}. {q["title"]}')
            lines.append(answers_map.get(q['id'], '') or '(no answer)')
            lines.append('')
        msg = MIMEMultipart()
        msg['From']    = f'{SENDER_NAME} <{SMTP_USER}>'
        msg['To']      = student_email
        msg['Subject'] = 'Your submission has been received'
        msg.attach(MIMEText('\n'.join(lines), 'plain'))
        with _smtp() as srv:
            srv.send_message(msg)
    except Exception as exc:
        print(f'[Confirmation email] {exc}')


# ─── Excel builder ─────────────────────────────────────────────────────────────

def _parse_feedback(raw):
    try:
        d = json.loads(raw) if raw else {}
    except Exception:
        return {'_raw': raw or ''}

    safety     = d.get('safety', '')
    safety_str = ('✓ CLEARED' if safety == 'CLEARED'
                  else ('✗ NOT CLEARED: ' + d.get('safety_reason', '')) if safety else '')

    na = {'safety': safety_str or '—', 'pass_rate': 'N/A', 'avg_time': 'N/A',
          'max_time': 'N/A', 'min_time': 'N/A', 'peak_mem': 'N/A',
          'wrong': 'N/A', 'exceptions': 'N/A'}

    if safety == 'NOT CLEARED':
        return na

    if d.get('note'):
        return {**na, 'safety': safety_str or '✓ CLEARED', 'pass_rate': d['note']}

    total  = d.get('total', 0)
    passed = d.get('passed', 0)

    if d.get('error') and total == 0:
        return {**na, 'safety': safety_str or '✓ CLEARED',
                'pass_rate': f'Error: {d["error"]}'}

    return {
        'safety':     safety_str or '✓ CLEARED',
        'pass_rate':  f"{d.get('pass_rate', 0):.1f}%  ({passed}/{total})",
        'avg_time':   f"{d.get('avg_time_ms', 0):.3f} ms",
        'max_time':   f"{d.get('max_time_ms', 0):.3f} ms",
        'min_time':   f"{d.get('min_time_ms', 0):.3f} ms",
        'peak_mem':   f"{d.get('peak_memory_mb', 0):.4f} MB",
        'wrong':      str(d.get('wrong_answers', 0)),
        'exceptions': str(d.get('exceptions', 0)),
    }


_CODE_COLS = [
    ('— Safety',        18, 'safety'),
    ('— Pass Rate',     22, 'pass_rate'),
    ('— Avg Time',      16, 'avg_time'),
    ('— Max Time',      16, 'max_time'),
    ('— Min Time',      16, 'min_time'),
    ('— Peak Memory',   16, 'peak_mem'),
    ('— Wrong Answers', 16, 'wrong'),
    ('— Exceptions',    14, 'exceptions'),
]


def build_excel(submission_ids):
    if not submission_ids:
        return None
    placeholders = ','.join('?' * len(submission_ids))
    conn = get_db()
    rows = conn.execute(f'''
        SELECT s.id, s.student_email, s.submitted_at,
               a.question_id, a.answer_text, a.claude_feedback,
               q.title, q.order_num, q.template
        FROM submissions s
        JOIN answers a ON a.submission_id = s.id
        JOIN questions q ON q.id = a.question_id
        WHERE s.id IN ({placeholders})
        ORDER BY s.student_email, s.id, q.order_num, q.id
    ''', submission_ids).fetchall()
    conn.close()
    if not rows:
        return None

    sub_map = OrderedDict()
    q_meta  = OrderedDict()

    for row in rows:
        sid = row['id']
        if sid not in sub_map:
            sub_map[sid] = {'email': row['student_email'], 'at': row['submitted_at'],
                            'answers': {}, 'feedback': {}}
        sub_map[sid]['answers'][row['question_id']]  = row['answer_text'] or ''
        sub_map[sid]['feedback'][row['question_id']] = row['claude_feedback'] or ''
        if row['question_id'] not in q_meta:
            q_meta[row['question_id']] = {'title': row['title'], 'is_code': bool(row['template'])}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Submissions'

    blue   = PatternFill('solid', fgColor='1D4ED8')
    green  = PatternFill('solid', fgColor='15803D')
    purple = PatternFill('solid', fgColor='6D28D9')
    wbold  = Font(color='FFFFFF', bold=True)
    wrap   = Alignment(wrap_text=True, vertical='top')

    header_spec = [('#', blue, 5), ('Email', blue, 32), ('Submitted At', blue, 22)]
    for qid, meta in q_meta.items():
        header_spec.append((meta['title'], blue, 45))
        if meta['is_code']:
            for suffix, width, _ in _CODE_COLS:
                fill = green if 'Pass Rate' in suffix else purple
                header_spec.append((f'{meta["title"]} {suffix}', fill, width))

    for ci, (label, fill, _) in enumerate(header_spec, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = wbold; cell.fill = fill; cell.alignment = wrap

    for ci, (_, _, width) in enumerate(header_spec, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = width

    for ri, (sid, data) in enumerate(sub_map.items(), 2):
        col = 1
        ws.cell(row=ri, column=col, value=ri - 1);          col += 1
        ws.cell(row=ri, column=col, value=data['email']);    col += 1
        ws.cell(row=ri, column=col, value=data['at']);       col += 1
        for qid, meta in q_meta.items():
            ws.cell(row=ri, column=col, value=data['answers'].get(qid, '')).alignment = wrap
            col += 1
            if meta['is_code']:
                fb = _parse_feedback(data['feedback'].get(qid, ''))
                for _, _, key in _CODE_COLS:
                    ws.cell(row=ri, column=col, value=fb.get(key, '')).alignment = wrap
                    col += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Send report ───────────────────────────────────────────────────────────────

def send_report():
    """Analyze latest submission per student, build Excel, email professor."""
    if not all([PROFESSOR_EMAIL, SMTP_USER, SMTP_PASS]):
        return False, 'Email not configured — check PROFESSOR_EMAIL, SMTP_USER, SMTP_PASS in .env'

    conn = get_db()
    latest = conn.execute('''
        SELECT student_email, MAX(id) AS latest_id
        FROM submissions
        WHERE sent = 0
        GROUP BY student_email
    ''').fetchall()

    if not latest:
        conn.close()
        return False, 'No new submissions to send'

    latest_ids    = [row['latest_id'] for row in latest]
    questions_list = [dict(q) for q in conn.execute(
        'SELECT * FROM questions ORDER BY order_num, id'
    ).fetchall()]
    conn.close()

    # Analyze all latest submissions in parallel (up to 5 at a time)
    with ThreadPoolExecutor(max_workers=min(5, len(latest_ids))) as ex:
        futures = [ex.submit(_analyze_one_submission, sid, questions_list) for sid in latest_ids]
        for f in futures:
            try:
                f.result()
            except Exception as e:
                print(f'[Analysis] {e}')

    excel_bytes = build_excel(latest_ids)
    if not excel_bytes:
        return False, 'Could not build Excel'

    today = date.today().strftime('%Y-%m-%d')
    msg = MIMEMultipart()
    msg['From']    = f'{SENDER_NAME} <{SMTP_USER}>'
    msg['To']      = PROFESSOR_EMAIL
    msg['Subject'] = f'Student Submissions — {today}'
    msg.attach(MIMEText(
        f'{len(latest_ids)} student(s) included (latest submission per student only).\n'
        f'Each code question shows: safety status, pass rate %, avg/max/min time, peak memory, error counts.',
        'plain'
    ))
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(excel_bytes)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="submissions_{today}.xlsx"')
    msg.attach(part)

    try:
        with _smtp() as srv:
            srv.send_message(msg)
        conn = get_db()
        conn.execute('UPDATE submissions SET sent=1 WHERE sent=0')
        conn.commit()
        conn.close()
        return True, f'Sent {len(latest_ids)} student submission(s) to {PROFESSOR_EMAIL}'
    except Exception as exc:
        return False, f'Email failed: {exc}'


# ─── Startup ───────────────────────────────────────────────────────────────────

init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=False)
